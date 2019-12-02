from openpyxl import load_workbook
from collections import defaultdict


wb = load_workbook('data.xlsx')
sheet = wb['Лист1']

waves = []

floors = []

wave_workers = dict()

route_floor = dict()

box_floor = dict()

route_num = defaultdict(int)

box_num = defaultdict(int)

num_boxes = defaultdict(set)

box_cell_d = defaultdict(list)

route_cell_d = defaultdict(list)


route_cell = defaultdict(list)
id_route = defaultdict(lambda: route_cell.copy())
wave_id = defaultdict(lambda: id_route.copy())

id_num = defaultdict(list)
wave_id_num = defaultdict(lambda: id_num.copy())

box_cell = defaultdict(set)
floor_box_cell = defaultdict(lambda: box_cell.copy())
wave_floor_box_cell = defaultdict(lambda: floor_box_cell.copy())

box_cell_lst = defaultdict(list)
route_box_cell = defaultdict(lambda: box_cell_lst.copy())

floor_route = defaultdict(list)
wave_floor_route = defaultdict(lambda: floor_route.copy())

route_box = defaultdict(list)
floor_route_box = defaultdict(lambda: route_box.copy())
wave_floor_route_box = defaultdict(lambda: floor_route_box.copy())

floor_id = defaultdict(list)
wave_floor_id = defaultdict(lambda: floor_id.copy())

idroute = defaultdict(list)
waveid = defaultdict(lambda: idroute.copy())
floor_wave_id_route = defaultdict(lambda: waveid.copy())

order_volume = defaultdict(list)
floor_order_volume = defaultdict(lambda: order_volume.copy())
wave_floor_order_volume = defaultdict(lambda: floor_order_volume.copy())

for i in range(1, 55432):

    wave = sheet.cell(row=i, column=2).value
    order = sheet.cell(row=i, column=3).value
    num = sheet.cell(row=i, column=5).value
    volume = sheet.cell(row=i,column=6).value
    box = sheet.cell(row=i, column=7).value
    id = sheet.cell(row=i, column=9).value
    route = sheet.cell(row=i, column=10).value
    cell = sheet.cell(row=i, column=12).value
    floor = cell[0]

    box_num[box] += num

    box_floor[box] = floor

    route_num[route] += num

    num_boxes[route].add(box)

    route_floor[route] = floor

    box_cell_d[box].append(cell)

    route_cell_d[route].append(cell)

    wave_id_num[wave][id].append(num)

    route_box_cell[route][box].append(int(cell[1:3]))

    wave_floor_order_volume[wave][floor][order].append([volume, cell])

    wave_floor_box_cell[wave][floor][box].add((int(cell[1:3]) - 1) // 2)

    wave_id[wave][id][route].append(list(map(int, cell[1:-2].split("-"))))



    if floor not in floors:
        floors.append(floor)

    if wave not in waves:
        waves.append(wave)

    if route not in wave_floor_route[wave][floor]:
        wave_floor_route[wave][floor].append(route)

    if id not in wave_floor_id[wave][floor]:
        wave_floor_id[wave][floor].append(id)

    if route not in floor_wave_id_route[floor][wave][id]:
        floor_wave_id_route[floor][wave][id].append(route)

    if box not in wave_floor_route_box[wave][floor][route]:
        wave_floor_route_box[wave][floor][route].append(box)


for j in wave_id.keys():
    wave_workers[j] = len(wave_id[j].keys())

floors.sort()



