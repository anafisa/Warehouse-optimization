from pprint import pprint
from openpyxl import load_workbook
from collections import defaultdict


wb = load_workbook('data.xlsx')
sheet = wb['Лист1']

waves = []

floors = []

wave_workers = dict()

route_num = defaultdict(lambda: 0)

route_cell = defaultdict(lambda: [])
id_route = defaultdict(lambda: route_cell.copy())
wave_id = defaultdict(lambda: id_route.copy())

id_num = defaultdict(lambda: [])
wave_id_num = defaultdict(lambda: id_num.copy())

route_cell_d = defaultdict(lambda: [])

floor_route = defaultdict(lambda: [])
wave_floor_route = defaultdict(lambda: floor_route.copy())

floor_id = defaultdict(lambda: [])
wave_floor_id = defaultdict(lambda: floor_id.copy())

idroute = defaultdict(lambda: [])
waveid = defaultdict(lambda: idroute.copy())
floor_wave_id_route = defaultdict(lambda: waveid.copy())


for i in range(1, 55432):
    wave = sheet.cell(row=i, column=2).value
    id = sheet.cell(row=i, column=9).value
    route = sheet.cell(row=i, column=10).value
    cell = sheet.cell(row=i, column=12).value
    num = sheet.cell(row=i, column=5).value
    floor = cell[0]

    wave_id[wave][id][route].append(list(map(int, cell[1:-2].split("-"))))  # {wave1:{id1:[{route1:[cell1,cell2]}], id2}

    wave_id_num[wave][id].append(num)  # {wave1:{id1:[num1,num2],id2} - number of goods which picks each person (id)

    route_cell_d[route].append(cell)  # {route1:[cell1, cell2], route2:[]}

    route_num[route] += num

    if floor not in floors:
        floors.append(floor)

    if wave not in waves:
        waves.append(wave)

    if route not in wave_floor_route[wave][floor]:
        wave_floor_route[wave][floor].append(route)  # {wave1:{'1':[r1,r2]}, wave2:}

    if id not in wave_floor_id[wave][floor]:
        wave_floor_id[wave][floor].append(id)  # wave1:{'1':[id1,id2]}, wave2:}

    if route not in floor_wave_id_route[floor][wave][id]:
        floor_wave_id_route[floor][wave][id].append(route)

for j in wave_id.keys():
    wave_workers[j] = len(wave_id[j].keys())

floors.sort()


